import pandas as pd
from functools import reduce
from openpyxl import load_workbook
from datetime import datetime as dt
import os
import time

start_time = time.time()

os.chdir('C:/Users/Jisnu/Desktop/HouseHoldData')

def append_df_to_excel(filename, df,sheet_name='Sheet1',startrow=None, truncate_sheet=False, **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Returns: None
    """
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0
    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs, header = False, index = False)
    writer.save()


ALR = pd.read_excel('ALR.xlsx')
HH = pd.read_excel('Allstate HouseHold Data.xlsx')
IV = pd.read_excel('IVantage Policy Book.xlsx')

HH.drop_duplicates(keep = 'first', inplace = True)
HH['Policy Number'] = [s.lstrip("0") for s in HH['policy_number']]
ALR['Annual Premium Amount'].fillna(0, inplace = True)
IV['Written Permium'].fillna(0, inplace = True)
ALR = ALR.reindex(columns = ['Policy Number','Annual Premium Amount'])
IV = IV.reindex(columns = ['Policy Number','Written Permium'])

ALR['Policy Number'] = ALR['Policy Number'].astype(str)
ALR['Policy Number'] = [s.lstrip("0") for s in ALR['Policy Number']]

#HH.loc[HH.Policy Number.isin(ALR['Policy Number']),['written_premium']]== ALR.loc[ALR['Policy Number'].isin(HH['Policy Number']),['Annual Premium Amount']].values
#HouseHoldData = HH.merge(ALR,on='Policy Number').merge(IV,on='Policy Number')
#HouseHoldData = pd.merge(pd.merge(HH,ALR,on='Policy Number'),IV,on='Policy Number')

df = pd.merge(HH, IV, on = 'Policy Number', how = 'left')
HouseHoldData = pd.merge(df, ALR, on = 'Policy Number', how = 'left')

HouseHoldData['Calculated Premium'] = HouseHoldData['Written Permium'].fillna(0) + HouseHoldData['written_premium']
HouseHoldData.drop(['policy_number','Written Permium','written_premium'], axis = 1, inplace = True)

#append_df_to_excel('test.xlsx',HouseHoldData)
HouseHoldData.to_excel('runHouseHoldAggregation.xlsx', index = False)